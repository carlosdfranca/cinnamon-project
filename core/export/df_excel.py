from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== Estilos comuns =====
bold = Font(bold=True)
italic = Font(italic=True)
left = Alignment(horizontal="left")
right = Alignment(horizontal="right")
center = Alignment(horizontal="center")
indent2 = Alignment(horizontal="left", indent=2)
bottom_border = Border(bottom=Side(style="thin"))
underline_single = Border(bottom=Side(style="thin"))
underline_double = Border(bottom=Side(style="double"))
double_bottom_border = Border(bottom=Side(style="double"))


# ================================================================
# GUIA DPF (versão por datas)
# ================================================================
def criar_aba_dpf(
    wb, fundo, data_atual, data_anterior,
    dpf_tabela, pl_atual, pl_anterior,
    total_pl_passivo_atual, total_pl_passivo_anterior
):
    ws = wb.active
    ws.title = "DPF"
    ws.sheet_view.showGridLines = False

    nome_fundo = str(fundo.nome).upper()

    # =====================
    # Cabeçalho principal
    # =====================
    ws["A1"] = nome_fundo; ws["A1"].font = bold; ws["A1"].alignment = left
    ws["A2"] = f"CNPJ: {fundo.cnpj}"; ws["A2"].font = bold; ws["A2"].alignment = left
    ws["A3"] = f"Administrado por {fundo.empresa.nome}"; ws["A3"].alignment = left
    ws["A4"] = f"CNPJ: {fundo.empresa.cnpj or ''}"; ws["A4"].alignment = left

    ws.append([])
    ws["A6"] = "Demonstração da Posição Financeira"; ws["A6"].font = bold; ws["A6"].alignment = left
    ws["A7"] = f"Em {data_atual.strftime('%d/%m/%Y')} e {data_anterior.strftime('%d/%m/%Y')}"; ws["A7"].alignment = left
    ws["A8"] = "(Valores expressos em milhares de reais, exceto quando apresentado de outra forma)"
    ws["A8"].font = italic; ws["A8"].alignment = left
    ws.append([])

    # Inserir coluna vazia após a primeira
    ws.insert_cols(2)

    # =====================
    # Mapa de colunas
    # =====================
    COL = {
        "DESC": 1,
        "SEP_LEFT": 2,
        "Q_CUR": 3,
        "R_CUR": 4,
        "P_CUR": 5,
        "SEP_MID": 6,
        "Q_PRI": 7,
        "R_PRI": 8,
        "P_PRI": 9,
    }

    # Ajuste de larguras
    col_widths = {
        COL["DESC"]: 55,
        COL["SEP_LEFT"]: 3,
        COL["Q_CUR"]: 9,
        COL["R_CUR"]: 14,
        COL["P_CUR"]: 16,
        COL["SEP_MID"]: 3,
        COL["Q_PRI"]: 9,
        COL["R_PRI"]: 14,
        COL["P_PRI"]: 16,
    }
    for idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # =====================
    # Cabeçalho de datas
    # =====================
    row0 = ws.max_row + 2
    ws.cell(row=row0, column=COL["Q_CUR"],
            value=data_atual.strftime("%d/%m/%Y")).font = bold
    ws.cell(row=row0, column=COL["Q_CUR"]).alignment = center
    ws.merge_cells(start_row=row0, start_column=COL["Q_CUR"], end_row=row0, end_column=COL["P_CUR"])

    ws.cell(row=row0, column=COL["Q_PRI"],
            value=data_anterior.strftime("%d/%m/%Y")).font = bold
    ws.cell(row=row0, column=COL["Q_PRI"]).alignment = center
    ws.merge_cells(start_row=row0, start_column=COL["Q_PRI"], end_row=row0, end_column=COL["P_PRI"])

    # Subcabeçalhos
    row1 = row0 + 1
    headers = [
        (COL["DESC"], "Ativo", center),
        (COL["Q_CUR"], "Quant", center),
        (COL["R_CUR"], "R$", center),
        (COL["P_CUR"], "% sobre o patrimônio líquido",
         Alignment(horizontal="center", vertical="center", wrap_text=True)),
        (COL["Q_PRI"], "Quant", center),
        (COL["R_PRI"], "R$", center),
        (COL["P_PRI"], "% sobre o patrimônio líquido",
         Alignment(horizontal="center", vertical="center", wrap_text=True)),
    ]
    for col, val, align in headers:
        c = ws.cell(row=row1, column=col, value=val)
        c.font = bold
        c.alignment = align
        c.border = bottom_border

    # =====================
    # Helpers internos
    # =====================
    def _dash(v):
        try:
            return "-" if (v is None or float(v) == 0) else v
        except Exception:
            return v

    def _write_money(cell, value, underline=None, bold_=False):
        if value == "-" or value is None:
            cell.value = "-"
            cell.alignment = right
            cell.border = underline
            return
        cell.value = value
        cell.number_format = "#,##0_);(#,##0)"
        cell.alignment = right
        if bold_:
            cell.font = Font(bold=True)
        if underline:
            cell.border = underline

    def _write_percent(cell, value, underline=None, bold_=False):
        if value in (None, "-"):
            cell.value = "-"
            cell.alignment = right
            return
        cell.value = float(value)
        cell.number_format = "#,##0.00"
        cell.alignment = right
        if bold_:
            cell.font = Font(bold=True)
        if underline:
            cell.border = underline

    current_row = ws.max_row + 2

    def _add_linha(descricao, v_atual=None, p_atual=None, v_ant=None, p_ant=None,
                   bold_line=False, indent=False, underline_kind=None):
        nonlocal current_row
        r = current_row
        ws.cell(row=r, column=COL["DESC"], value=descricao)
        ws.cell(row=r, column=COL["DESC"]).alignment = indent2 if indent else left
        if bold_line:
            ws.cell(row=r, column=COL["DESC"]).font = Font(bold=True)
        ws.cell(row=r, column=COL["Q_CUR"], value="-").alignment = right
        _write_money(ws.cell(row=r, column=COL["R_CUR"]), _dash(v_atual), underline=underline_kind, bold_=bold_line)
        _write_percent(ws.cell(row=r, column=COL["P_CUR"]), p_atual, underline=underline_kind, bold_=bold_line)
        ws.cell(row=r, column=COL["SEP_MID"], value="")
        ws.cell(row=r, column=COL["Q_PRI"], value="-").alignment = right
        _write_money(ws.cell(row=r, column=COL["R_PRI"]), _dash(v_ant), underline=underline_kind, bold_=bold_line)
        _write_percent(ws.cell(row=r, column=COL["P_PRI"]), p_ant, underline=underline_kind, bold_=bold_line)
        current_row += 1

    def _perc_from(val, base):
        try:
            return round((float(val or 0) / float(base or 0)) * 100, 2) if base else 0.0
        except:
            return 0.0

    # =====================
    # Seções: ATIVO, PASSIVO, PL
    # =====================
    ativo = dpf_tabela["ATIVO"]
    for grupo_label, bloco in ativo.items():
        if grupo_label.startswith("TOTAL_"):
            continue
        soma_atual = bloco.get("SOMA", 0)
        soma_ant = bloco.get("SOMA_ANTERIOR", 0)
        p_atual = bloco.get("PERC_ATUAL", _perc_from(soma_atual, pl_atual))
        p_ant = bloco.get("PERC_ANTERIOR", _perc_from(soma_ant, pl_anterior))
        _add_linha(grupo_label, soma_atual, p_atual, soma_ant, p_ant, bold_line=True, underline_kind=underline_single)

        for sub, valores in bloco.items():
            if sub in ("SOMA", "SOMA_ANTERIOR"):
                continue
            if isinstance(valores, dict):
                v_at = valores.get("ATUAL", 0)
                v_an = valores.get("ANTERIOR", 0)
                p_at = valores.get("PERC_ATUAL", _perc_from(v_at, pl_atual))
                p_an = valores.get("PERC_ANTERIOR", _perc_from(v_an, pl_anterior))
                _add_linha(sub, v_at, p_at, v_an, p_an, indent=True)
        ws.append([]); current_row = ws.max_row + 2

    tot_ativo_at = ativo["TOTAL_ATIVO"]["ATUAL"]
    tot_ativo_an = ativo["TOTAL_ATIVO"]["ANTERIOR"]
    _add_linha("Total do ativo", tot_ativo_at, _perc_from(tot_ativo_at, pl_atual),
               tot_ativo_an, _perc_from(tot_ativo_an, pl_anterior),
               bold_line=True, underline_kind=underline_double)

    # ===== PASSIVO =====
    passivo = dpf_tabela["PASSIVO"]
    for grupo_label, bloco in passivo.items():
        if grupo_label.startswith("TOTAL_"):
            continue
        soma_atual = bloco.get("SOMA", 0)
        soma_ant = bloco.get("SOMA_ANTERIOR", 0)
        p_atual = bloco.get("PERC_ATUAL", _perc_from(soma_atual, pl_atual))
        p_ant = bloco.get("PERC_ANTERIOR", _perc_from(soma_ant, pl_anterior))
        _add_linha(grupo_label, soma_atual, p_atual, soma_ant, p_ant, bold_line=True, underline_kind=underline_single)

        for sub, valores in bloco.items():
            if sub in ("SOMA", "SOMA_ANTERIOR"):
                continue
            if isinstance(valores, dict):
                v_at = valores.get("ATUAL", 0)
                v_an = valores.get("ANTERIOR", 0)
                p_at = valores.get("PERC_ATUAL", _perc_from(v_at, pl_atual))
                p_an = valores.get("PERC_ANTERIOR", _perc_from(v_an, pl_anterior))
                _add_linha(sub, v_at, p_at, v_an, p_an, indent=True)
        ws.append([]); current_row = ws.max_row + 2

    tot_passivo_at = passivo["TOTAL_PASSIVO"]["ATUAL"]
    tot_passivo_an = passivo["TOTAL_PASSIVO"]["ANTERIOR"]
    _add_linha("Total do passivo", tot_passivo_at, _perc_from(tot_passivo_at, pl_atual),
               tot_passivo_an, _perc_from(tot_passivo_an, pl_anterior),
               bold_line=True, underline_kind=underline_double)

    ws.append([]); current_row = ws.max_row + 2

    # ===== PL =====
    _add_linha("Patrimônio líquido", pl_atual, 100.00, pl_anterior, 100.00,
               bold_line=True, underline_kind=underline_double)

    ws.append([]); current_row = ws.max_row + 2

    # ===== TOTAL PL + PASSIVO =====
    _add_linha("Total do patrimônio líquido e do passivo",
               total_pl_passivo_atual, _perc_from(total_pl_passivo_atual, pl_atual),
               total_pl_passivo_anterior, _perc_from(total_pl_passivo_anterior, pl_anterior),
               bold_line=True, underline_kind=underline_double)

    ws.append([]); current_row = ws.max_row + 2

    # Rodapé
    last_col = max(ws.max_column, 9)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
    cell = ws.cell(row=current_row, column=1)
    cell.value = "As notas explicativas são parte integrante das demonstrações financeiras."
    cell.font = Font(italic=True, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ================================================================
# GUIA DRE (versão por datas)
# ================================================================
def criar_aba_dre(wb, fundo, data_atual, data_anterior, dre_tabela, resultado_exercicio, resultado_exercicio_anterior):
    ws = wb.create_sheet(title="DRE")
    ws.sheet_view.showGridLines = False

    nome_fundo = str(fundo.nome).upper()

    # Cabeçalho
    ws["A1"] = nome_fundo; ws["A1"].font = bold; ws["A1"].alignment = left
    ws["A2"] = f"CNPJ: {fundo.cnpj}"; ws["A2"].font = bold; ws["A2"].alignment = left
    ws["A3"] = fundo.empresa.nome; ws["A3"].alignment = left
    ws["A4"] = f"CNPJ: {fundo.empresa.cnpj or ''}"; ws["A4"].alignment = left
    ws.append([])

    ws["A6"] = "Demonstração do Resultado do Exercício"
    ws["A6"].font = bold; ws["A6"].alignment = left
    ws["A7"] = f"Períodos findos em {data_atual.strftime('%d/%m/%Y')} e {data_anterior.strftime('%d/%m/%Y')}"
    ws["A7"].font = bold; ws["A7"].alignment = left
    ws["A8"] = "(Valores expressos em milhares de reais)"
    ws["A8"].font = italic; ws["A8"].alignment = left
    ws.append([])

    # Cabeçalho das colunas
    ws.insert_cols(3)
    ws.append(["", data_atual.strftime("%d/%m/%Y"), "", data_anterior.strftime("%d/%m/%Y")])
    row_header = ws.max_row
    for col in (2, 4):
        c = ws.cell(row=row_header, column=col)
        c.alignment = right
        c.font = bold
        c.border = bottom_border

    # Linhas da DRE
    for grupo, dados in dre_tabela.items():
        ws.append([grupo, dados["SOMA"], "", dados["SOMA_ANTERIOR"]])
        row = ws.max_row
        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=1).alignment = left
        for col in (2, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = bold
            cell.alignment = right
            cell.number_format = "#,##0_);(#,##0)"
            cell.border = bottom_border

        for item, valores in dados.items():
            if item in ["SOMA", "SOMA_ANTERIOR"]:
                continue
            ws.append([item, valores["ATUAL"], "", valores["ANTERIOR"]])
            row = ws.max_row
            ws.cell(row=row, column=1).alignment = indent2
            for col in (2, 4):
                cell = ws.cell(row=row, column=col)
                cell.alignment = right
                cell.number_format = "#,##0_);(#,##0)"
        ws.append([])

    # Resultado final
    ws.append(["Resultado do exercício", resultado_exercicio, "", resultado_exercicio_anterior])
    row = ws.max_row
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=1).alignment = left
    for col in (2, 4):
        cell = ws.cell(row=row, column=col)
        cell.number_format = "#,##0_);(#,##0)"
        cell.font = bold
        cell.alignment = right
        cell.border = double_bottom_border

    # Ajuste de largura
    ws.insert_cols(1)
    for col_num, width in {1:3, 2:65, 3:12, 4:5, 5:12, 6:3}.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


# ================================================================
# GUIA DMPL (versão por datas)
# ================================================================
def criar_aba_dmpl(
    wb, fundo, data_atual, data_anterior,
    dados_dmpl, resultado_exercicio, resultado_exercicio_anterior,
    pl_atual, pl_anterior
):
    ws = wb.create_sheet(title="DMPL")
    ws.sheet_view.showGridLines = False

    nome_fundo = str(fundo.nome).upper()

    # =====================
    # Cabeçalho
    # =====================
    ws["A1"] = nome_fundo; ws["A1"].font = bold; ws["A1"].alignment = left
    ws["A2"] = f"CNPJ: {fundo.cnpj}"; ws["A2"].font = bold; ws["A2"].alignment = left
    ws["A3"] = f"Administrado por {fundo.empresa.nome}"; ws["A3"].alignment = left
    ws["A4"] = f"CNPJ: {fundo.empresa.cnpj or ''}"; ws["A4"].alignment = left
    ws.append([])

    ws["A6"] = "Demonstração das Mutações do Patrimônio Líquido"
    ws["A6"].font = bold; ws["A6"].alignment = left
    ws["A7"] = f"Períodos findos em {data_atual.strftime('%d/%m/%Y')} e {data_anterior.strftime('%d/%m/%Y')}"
    ws["A7"].font = bold; ws["A7"].alignment = left
    ws["A8"] = "(Valores expressos em milhares de reais, exceto o valor unitário da cota)"
    ws["A8"].font = italic; ws["A8"].alignment = left
    ws.append([])

    # =====================
    # Cabeçalho colunas
    # =====================
    ws.append(["Descrição", data_atual.strftime("%d/%m/%Y"), data_anterior.strftime("%d/%m/%Y")])
    row_header = ws.max_row
    for col in (2, 3):
        c = ws.cell(row=row_header, column=col)
        c.alignment = right
        c.font = bold
        c.border = bottom_border

    # =====================
    # PL inicial
    # =====================
    ws.append(["Patrimônio líquido no início do período",
               dados_dmpl["valor_primeiro"], dados_dmpl["valor_primeiro_ant"]])
    ws.append([f"Total de {dados_dmpl['qtd_cotas_inicio']} cotas a R$ {dados_dmpl['cota_inicio']}",
               dados_dmpl["valor_primeiro"], "-"])
    ws.append([f"Total de {dados_dmpl['qtd_cotas_inicio_ant']} cotas a R$ {dados_dmpl['cota_inicio_ant']}",
               "-", dados_dmpl["valor_primeiro_ant"]])
    ws.append([])

    # =====================
    # Emissão
    # =====================
    ws.append(["Emissão de cotas", "", ""])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([f"Total de {dados_dmpl['aplicacoes_qtd']} cotas",
               dados_dmpl["aplicacoes_valor"], dados_dmpl["aplicacoes_valor_ant"]])
    ws.append([])

    # =====================
    # Resgate
    # =====================
    ws.append(["Resgate de cotas", "", ""])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([f"Total de {dados_dmpl['resgates_qtd']} cotas",
               dados_dmpl["resgates_valor"], dados_dmpl["resgates_valor_ant"]])
    ws.append([])

    # =====================
    # PL antes do resultado
    # =====================
    ws.append([
        "Patrimônio líquido antes do resultado do período",
        dados_dmpl["pl_antes_resultado_periodo"],
        dados_dmpl["valor_primeiro_ant"]
    ])
    ws.append([])

    # =====================
    # Resultado
    # =====================
    ws.append(["Resultado do período", resultado_exercicio, resultado_exercicio_anterior])
    ws.append([])

    # =====================
    # PL final
    # =====================
    ws.append(["Patrimônio líquido no final do exercício/período", pl_atual, pl_anterior])
    ws.append([f"Total de {dados_dmpl['qtd_cotas_fim']} cotas a R$ {dados_dmpl['cota_fim']}",
               dados_dmpl["valor_ultimo"], "-"])
    ws.append([f"Total de {dados_dmpl['qtd_cotas_inicio']} cotas a R$ {dados_dmpl['cota_inicio']}",
               "-", dados_dmpl["valor_primeiro"]])
    ws.append([])

    # =====================
    # Observação final
    # =====================
    last_col = max(ws.max_column, 3)
    ws.merge_cells(
        start_row=ws.max_row + 1, start_column=1,
        end_row=ws.max_row + 1, end_column=last_col
    )
    cell = ws.cell(row=ws.max_row, column=1)
    cell.value = "As notas explicativas são parte integrante das demonstrações financeiras."
    cell.font = Font(italic=True, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # =====================
    # Ajuste de largura
    # =====================
    for col_num, width in {1: 65, 2: 15, 3: 15}.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


# ================================================================
# GUIA DFC (versão por datas)
# ================================================================
def criar_aba_dfc(wb, fundo, data_atual, data_anterior, dfc_tabela, variacao_atual, variacao_ant):
    ws = wb.create_sheet(title="DFC")
    ws.sheet_view.showGridLines = False

    nome_fundo = str(fundo.nome).upper()

    # =====================
    # Cabeçalho
    # =====================
    ws["A1"] = nome_fundo; ws["A1"].font = bold; ws["A1"].alignment = left
    ws["A2"] = f"CNPJ: {fundo.cnpj}"; ws["A2"].font = bold; ws["A2"].alignment = left
    ws["A3"] = f"Administrado por {fundo.empresa.nome}"; ws["A3"].alignment = left
    ws["A4"] = f"CNPJ: {fundo.empresa.cnpj or ''}"; ws["A4"].alignment = left
    ws.append([])

    ws["A6"] = "Demonstração dos Fluxos de Caixa – Método indireto"
    ws["A6"].font = bold; ws["A6"].alignment = left
    ws["A7"] = f"Períodos findos em {data_atual.strftime('%d/%m/%Y')} e {data_anterior.strftime('%d/%m/%Y')}"
    ws["A7"].font = bold; ws["A7"].alignment = left
    ws["A8"] = "(Valores expressos em milhares de reais)"
    ws["A8"].font = italic; ws["A8"].alignment = left
    ws.append([])

    # =====================
    # Cabeçalho colunas
    # =====================
    ws.append(["Descrição", data_atual.strftime("%d/%m/%Y"), data_anterior.strftime("%d/%m/%Y")])
    row_header = ws.max_row
    for col in (2, 3):
        c = ws.cell(row=row_header, column=col)
        c.alignment = right
        c.font = bold
        c.border = bottom_border

    # =====================
    # Helpers
    # =====================
    def _write_linha(descricao, atual=None, anterior=None, bold_=False, underline=None, indent=False):
        ws.append([descricao, atual, anterior])
        r = ws.max_row
        ws.cell(row=r, column=1).alignment = indent2 if indent else left
        if bold_:
            ws.cell(row=r, column=1).font = Font(bold=True)
        for c in (2, 3):
            cell = ws.cell(row=r, column=c)
            cell.alignment = right
            cell.number_format = "#,##0_);(#,##0)"
            if bold_:
                cell.font = Font(bold=True)
            if underline:
                cell.border = underline

    # =============================================================
    # BLOCO 1: ATIVIDADES OPERACIONAIS
    # =============================================================
    bloco_op = dfc_tabela.get("fluxo_operacionais", {})
    _write_linha(bloco_op.get("titulo", ""), bold_=True)
    ws.append([])

    # Resultado líquido do período
    resultado = bloco_op.get("resultado_liquido", {})
    _write_linha(
        resultado.get("titulo", ""),
        resultado.get("ATUAL"),
        resultado.get("ANTERIOR"),
        bold_=True,
        underline=underline_double,
        indent=True,
    )

    # ---- Ajustes
    ajustes = bloco_op.get("ajustes", {})
    if ajustes:
        ws.append([])
        _write_linha(ajustes.get("titulo", ""), bold_=True, indent=True)
        for chave, item in ajustes.items():
            if not isinstance(item, dict) or chave == "titulo":
                continue
            _write_linha(item.get("titulo", ""), item.get("ATUAL"), item.get("ANTERIOR"), indent=True)
        ws.append([])

    # ---- Linhas subsequentes (variações)
    for chave in ["aumento_dc", "aumento_receber", "reducao_pagar", "caixa_operacional"]:
        item = bloco_op.get(chave, {})
        if not item:
            continue
        bold_line = "caixa" in chave
        underline_kind = underline_double if "caixa" in chave else None
        _write_linha(
            item.get("titulo", ""),
            item.get("ATUAL"),
            item.get("ANTERIOR"),
            bold_=bold_line,
            underline=underline_kind,
            indent=True,
        )

    ws.append([])

    # =============================================================
    # BLOCO 2: ATIVIDADES DE FINANCIAMENTO
    # =============================================================
    bloco_fin = dfc_tabela.get("fluxo_financiamento", {})
    _write_linha(bloco_fin.get("titulo", ""), bold_=True)
    for chave, item in bloco_fin.items():
        if not isinstance(item, dict) or chave == "titulo":
            continue
        bold_line = "caixa" in chave
        underline_kind = underline_double if "caixa" in chave else None
        _write_linha(
            item.get("titulo", ""),
            item.get("ATUAL"),
            item.get("ANTERIOR"),
            bold_=bold_line,
            underline=underline_kind,
            indent=True,
        )
    ws.append([])

    # =============================================================
    # BLOCO 3: VARIAÇÃO E CAIXA FINAL
    # =============================================================

    # Variação no caixa e equivalentes (linha dupla)
    item_var = dfc_tabela.get("variacao_caixa", {})
    if item_var:
        _write_linha(
            item_var.get("titulo", ""),
            item_var.get("ATUAL"),
            item_var.get("ANTERIOR"),
            bold_=True,
            underline=underline_double,
        )

    # Linha em branco antes do caixa inicial/final
    ws.append([])

    # Caixa início
    item_ini = dfc_tabela.get("caixa_inicio", {})
    if item_ini:
        _write_linha(item_ini.get("titulo", ""), item_ini.get("ATUAL"), item_ini.get("ANTERIOR"), bold_=True)

    # Caixa final
    item_fim = dfc_tabela.get("caixa_final", {})
    if item_fim:
        _write_linha(item_fim.get("titulo", ""), item_fim.get("ATUAL"), item_fim.get("ANTERIOR"), bold_=True)

    # Linha em branco antes das notas
    ws.append([])

    # Observação final
    last_col = max(ws.max_column, 3)
    ws.merge_cells(start_row=ws.max_row + 1, start_column=1, end_row=ws.max_row + 1, end_column=last_col)
    cell = ws.cell(row=ws.max_row, column=1)
    cell.value = "As notas explicativas são parte integrante das demonstrações financeiras."
    cell.font = Font(italic=True, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Largura das colunas
    for col_num, width in {1: 70, 2: 15, 3: 15}.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
